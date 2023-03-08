#!usr/env/python3
# -*- coding: utf-8 -*-

"""
# File       : department.py
# Time       : 2023/2/28 13:51
# Author     : Ron
# version    : python 3.9
# software   : Pycharm
# Description：部门信息管理
"""

from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from app01.utils.pagination import Pagination

# Create your views here.


def depart_list(request):
    """部门列表"""

    # 去数据库中获取所有部门列表
    depart_info = models.Department.objects.all()
    depart_obj = Pagination(request, depart_info, page_size=4)
    context = {
        "query_set": depart_obj.page_queryset,
        "page_string": depart_obj.html(),
    }
    return render(request, "depart_list.html", context)


def depart_add(request):
    """添加部门"""

    # 如果操作请求是Get
    if request.method == "GET":
        return render(request, "depart_add.html")
    # 如果操作请求是Post
    # 获取用户提交的数据
    title = request.POST.get("title")
    # 将获取到的数据写入数据库
    models.Department.objects.create(title=title)
    # 页面重定向
    return redirect("/depart/list/")


def depart_del(request, nid):
    """删除功能"""
    # http://127.0.0.1:8000/depart/del/?nid=1
    models.Department.objects.filter(id=nid).delete()

    return redirect("/depart/list/")


def depart_edit(request, nid):
    """修改功能"""
    if request.method == "GET":
        depart_info = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"depart_info": depart_info})

    # 获取用户数据
    title = request.POST.get("title")
    # 根据ID找到数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)

    return redirect("/depart/list/")


def depart_multi(request):
    """ 批量上传 """
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
